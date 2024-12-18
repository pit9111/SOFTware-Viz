import os
import json
from tqdm import tqdm
import xml.etree.ElementTree as ET
import requests
import time
from openpyxl import load_workbook
from Utils.TEI_to_JSON import transformer_TEI_JSON

def duplicates_JSON(lst):
    seen = set()
    duplicates = []

    for item in lst:
        item_hashable = str(item)
        if item_hashable in seen:
            duplicates.append(item)
        else:
            seen.add(item_hashable)

    return duplicates

def insert_json_db(data_path_json,data_path_xml,db):
    software_document = []
    list_errors = []

    workbook = load_workbook(filename='./app/static/data/Logiciels_Blacklist_et_autres_remarques.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    blacklist = []
    for row in data[1:]:
        blacklist.append(row[0])

    if db.hasCollection('documents'):
        documents_collection = db['documents']
    else:
        db.createCollection('Collection', name='documents')
        documents_collection = db['documents']

    if db.hasCollection('softwares'):
        softwares_collection = db['softwares']
    else:
        db.createCollection('Collection', name='softwares')
        softwares_collection = db['softwares']

    if db.hasCollection('references'):
        references_collection = db['references']
    else:
        db.createCollection('Collection', name='references')
        references_collection = db['references']

    if db.hasCollection('edge_software'):
        doc_soft_edge = db['edge_software']
    else:
        db.createCollection('Edges', name='edge_software')
        doc_soft_edge = db['edge_software']

    if db.hasCollection('edge_reference'):
        doc_ref_edge = db['edge_reference']
    else:
        db.createCollection('Edges', name='edge_reference')
        doc_ref_edge = db['edge_reference']


    data_json_files = os.listdir(data_path_json)
    data_xml_list = os.listdir(data_path_xml)
    files_list_registered = db.AQLQuery('FOR hal_id in documents RETURN hal_id.file_hal_id', rawResults=True)

    for data_file_xml in tqdm(data_xml_list):
        file_path = f'{data_path_xml}/{data_file_xml}'
        file_name = os.path.basename(file_path)
        while "." in file_name:
            file_name, extension = os.path.splitext(file_name)

        if file_name in files_list_registered:
            break
        time.sleep(0.2)
        url = "https://api.archives-ouvertes.fr/search/"
        params = {
            "q": f"halId_id:{file_name}",
            "rows": 10,
            "fl": "citationFull_s"
        }

        response = requests.get(url, params=params)
        if response.status_code == 200:
            data = response.json()
            try:
                citations = data["response"]["docs"][0]["citationFull_s"]
            except IndexError:
                citations = None
        else:
            print("Error:", response.status_code)

        with open(file_path, 'r', encoding='utf-8') as xml_file:
            data_json_get_document = {}
            tree = ET.parse(xml_file)
            root = tree.getroot()
            ns = {"tei": "http://www.tei-c.org/ns/1.0"}

            title = tree.find(".//tei:listBibl//tei:titleStmt//tei:title", ns)
            title = title.text

            structures = tree.find(".//tei:back//tei:listOrg", ns)
            list_org = []
            for org in structures:
                if org.attrib['type'] in ['regrouplaboratory','regroupinstitution','institution']:
                    for org_child_tags in list(org):
                        try:
                            if org_child_tags.attrib['type'] == 'acronym':
                                structure_name = org_child_tags.text
                                structure_name_sanitized = structure_name.rstrip()
                                list_org.append(structure_name_sanitized)
                        except KeyError:
                            continue
            data_json_get_document['structures'] = list_org

            doc_type = tree.findall(".//tei:listBibl//tei:biblFull//tei:profileDesc//tei:textClass//tei:classCode", ns)
            for tag in doc_type:
                if tag.attrib.get('n') == 'COMM':
                    production_date = tree.find(".//tei:listBibl//tei:biblFull//tei:sourceDesc//tei:biblStruct//tei:monogr//tei:meeting//tei:date[@type='start']", ns)
                    if production_date is not None:
                        data_json_get_document['date'] = production_date.text[:4]
                else:
                    production_date = tree.find(
                        ".//tei:listBibl//tei:biblFull//tei:sourceDesc//tei:biblStruct//tei:monogr//tei:imprint//tei:date[@type='datePub']",
                        ns)
                    if production_date is not None:
                        data_json_get_document['date'] = production_date.text[:4]
                    else:
                        production_date = tree.find(
                            ".//tei:listBibl//tei:biblFull//tei:editionStmt//tei:edition[@type='current']//tei:date[@type='whenProduced']",
                            ns)
                        if production_date is not None:
                            data_json_get_document['date'] = production_date.text[:4]
                        else:
                            print(f'problem : {file_name}')
            author_list = tree.findall(".//tei:listBibl//tei:titleStmt//tei:author", ns)
            list_author = []
            for elm in author_list:
                author = {}
                persName = elm.find("{http://www.tei-c.org/ns/1.0}persName")
                author['role'] = elm.attrib['role']
                for name in persName:
                    author[name.tag.split('}')[1]] =  name.text
                list_author.append(author)

            abstract = tree.find(".//{http://www.tei-c.org/ns/1.0}abstract")
            if abstract:
                tag_text = list(abstract)[0]
                if tag_text.tag == '{http://www.tei-c.org/ns/1.0}p':
                    abstract = "".join(tag_text.itertext())
                    data_json_get_document['abstract'] = ['HAL' , abstract]
                if tag_text.tag == '{http://www.tei-c.org/ns/1.0}div':
                    for p_tag in list(tag_text):
                        text = "".join(p_tag.itertext())
                    data_json_get_document['abstract'] = ['GROBID' , abstract]

            data_json_get_document['file_hal_id'] = file_name
            data_json_get_document['citation'] = citations
            data_json_get_document['title'] = title
            data_json_get_document['author'] = list_author

            document_document = documents_collection.createDocument(data_json_get_document)
            document_document.save()
            if f"{file_name}.software.json" in data_json_files:
                with open(f'{data_path_json}/{file_name}.software.json', 'r') as json_file:
                    data_json = json.load(json_file)
                    data_json_get_mentions = data_json.get("mentions")

                    # Remove duplicates
                    for elm in duplicates_JSON(data_json_get_mentions):
                        data_json_get_mentions.remove(elm)

                    # Process each mention
                    for mention in data_json_get_mentions:
                        if mention['software-name']['normalizedForm'] not in blacklist:
                            mention['software_name'] = mention.pop('software-name')
                            mention['software_type'] = mention.pop('software-type')
                            software_document = softwares_collection.createDocument(mention)
                            software_document.save()

                            # Create edge from document to software
                            edge = doc_soft_edge.createEdge()
                            edge['_from'] = document_document._id
                            edge['_to'] = software_document._id
                            edge.save()

                    # Process each reference
                    data_json_get_references = data_json.get("references")
                    for reference in data_json_get_references:
                        result_json = []
                        try:
                            result_json,error = transformer_TEI_JSON(reference['tei'])
                            if len(error) > 0:
                                list_errors.append(error, reference['tei'])
                        except Exception as e:
                            print(f"Error during the transformation from XML to JSON: {e}")
                        if result_json:
                            reference['json'] = result_json
                        references_document = references_collection.createDocument(reference)
                        references_document.save()

                        # Create edge from document to reference
                        edge = doc_ref_edge.createEdge()
                        edge['_from'] = document_document._id
                        edge['_to'] = references_document._id
                        edge.save()
                # Define the AQL query to fetch software names and their counts
                query = f"""
                FOR doc IN edge_software
                    FILTER doc._from == "{document_document._id}"
                    LET software = DOCUMENT(doc._to)
                    COLLECT softwareName = software.software_name.normalizedForm WITH COUNT INTO count
                    RETURN {{ softwareName, count }}
                """

                # Execute the AQL query to get software names and counts
                all_software_dict = db.AQLQuery(query, rawResults=True)

                # Fetch distinct software names
                distinct_query = f"""
                FOR doc IN edge_software
                    FILTER doc._from == "{document_document._id}"
                    LET software = DOCUMENT(doc._to)
                    RETURN DISTINCT software.software_name.normalizedForm
                """
                software_name_list = db.AQLQuery(distinct_query, rawResults=True)

                # Convert software names and counts to a dictionary
                dict_software = {software_dict['softwareName']: software_dict['count'] for software_dict in
                                 all_software_dict}

                # Process software names containing hyphens and update as needed
                for software_name in software_name_list:
                    if "-" in software_name:
                        software_name_cleaned = software_name.replace('-', '')
                        if software_name_cleaned in dict_software and dict_software[software_name] < dict_software[
                            software_name_cleaned]:
                            # Example: Update software name in the database
                            update_query = f"""
                            FOR doc IN edge_software
                                FILTER doc._from == "{document_document._id}"
                                LET software = DOCUMENT(doc._to)
                                FILTER software.software_name.normalizedForm == "{software_name}"
                                UPDATE software WITH {{ software_name: {{ normalizedForm: "{software_name_cleaned}" }} }} IN softwares
                            """
                            db.AQLQuery(update_query)
    if len(list_errors) > 0:
       print(list_errors)